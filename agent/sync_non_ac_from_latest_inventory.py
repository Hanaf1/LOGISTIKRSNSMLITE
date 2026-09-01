#!/usr/bin/env python3
"""Build a deterministic replacement for the latest non-AC inventory.

The current AC rows (item 021300/021301) and all Asuransi rows (unit 04) are
kept untouched. Historical snapshots embedded in several workbook sheets are
excluded. Every remaining source row becomes one asset group and its quantity
is expanded into individually numbered assets.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
import pymysql

import import_inventaris_non_medis_20260813 as legacy


DEFAULT_WORKBOOK = Path(r"C:\Users\ASUS\Downloads\LIST INVENTARIS_NON MEDIS.xlsx")
DEFAULT_REGISTRATION = Path(
    r"C:\Users\ASUS\Downloads\registrasi_aset_disesuaikan_dengan_list_inventaris.xlsx"
)
DEFAULT_REFERENCE = Path(
    r"C:\Users\ASUS\Downloads\FINAL_Import_Registrasi_Aset_Dengan_Referensi.xlsx"
)
AC_CODES = {"021300", "021301"}
PRESERVED_UNIT = "04"
SYNC_USER = "sync_non_ac_list_20260825"

# Rows whose LIST number is absent or demonstrably malformed. These mappings
# follow the existing item taxonomy and the actual item description.
ITEM_OVERRIDES = {
    ("Madinah 2", 36): "050601",       # Tempat Obat Pasien RI
    ("igd", 35): "050502",             # Container Box 25
    ("Laborat", 15): "032300",         # Rak/peniris alat stainless
    ("Laborat", 17): "051400",         # Keranjang resep/dokumen kecil
    ("Laborat", 19): "051700",         # Laci plastik susun
    ("Laborat", 38): "030103",         # Kursi tunggu panjang
    ("Laborat", 40): "050514",         # Container Box 45
    ("kasir", 93): "034400",           # Almari loker
    ("Madinah 4", 51): "041800",       # Meja nurse station
    ("Madinah 4", 64): "030601",       # Almari alkes
    ("CSSD", 6): "033100",             # Bak stainless
    ("CSSD", 7): "021600",             # Sealer plastik / press pouches
    ("gizi", 182): "051700",           # Laci plastik susun
    ("gizi", 193): "101300",           # Papan tulis akrilik
    ("RM", 62): "082200",              # Slempang mobile JKN
    ("OK", 13): "033100",              # Bak stainless
    ("Madinah 3", 21): "020300",       # Set komputer
    ("Madinah 3", 36): "050601",       # Tempat Obat Pasien RI
    ("Selasar Lt.1 Gd lama", 5): "030103",
    ("Selasar Lt.1 Gd lama", 13): "025801",
    ("TPS B3", 6): "063802",           # Lampu UV
    ("Arofah", 37): "030102",          # Kursi susun
    ("Server", 4): "020300",           # Komputer
}

# New item masters needed by the newest LIST. Codes use an existing group and
# kind, so they remain compatible with the current inventory numbering scheme.
REQUIRED_ITEM_MASTERS = {
    "032300": {
        "nama": "Peniris Alat Stainless", "kelompok": "03", "jenis": "23", "barang": "00",
        "nama_kelompok": "Mebelair Bahan Logam", "nama_jenis": "Rak peniris alat stainless",
    },
    "030601": {
        "nama": "Almari Alkes", "kelompok": "03", "jenis": "06", "barang": "01",
        "nama_kelompok": "Mebelair Bahan Logam", "nama_jenis": "Almari dokumen",
    },
    "063802": {
        "nama": "Lampu UV", "kelompok": "06", "jenis": "38", "barang": "02",
        "nama_kelompok": "APD dan Alat Rumah Tangga Lain-Lain", "nama_jenis": "Lampu",
    },
    "082200": {
        "nama": "Slempang Mobile JKN", "kelompok": "08", "jenis": "22", "barang": "00",
        "nama_kelompok": "Linen", "nama_jenis": "Slempang mobile JKN",
    },
}


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def digits(value) -> str:
    return re.sub(r"\D", "", clean(value))


def normalized(value) -> str:
    value = unicodedata.normalize("NFKD", clean(value)).casefold()
    return " ".join(re.findall(r"[a-z0-9]+", value))


def unit_code(number: int) -> str:
    return f"{number:02d}" if number < 100 else str(number)


def explicit_location_unit(location: str) -> int | None:
    matches = re.findall(r"\(\s*'?\s*(\d{1,3})\s*\)", clean(location))
    if not matches:
        return None
    number = int(matches[-1])
    return number if 1 <= number <= 142 else None


def source_sequence(value: str) -> int:
    number = digits(value)
    if len(number) < 2:
        return 1
    result = int(number[-3:]) if len(number) >= 3 else int(number[-2:])
    return result if 1 <= result <= 999 else 1


def is_historical(row: dict) -> bool:
    sheet = row["sheet"]
    number = int(row["row"])
    return (
        (sheet == "radiologi" and number < 79)
        or (sheet == "logistik" and number < 224)
        or (sheet == "Kebersihan" and number < 40)
        or (sheet == "RM" and number < 40)
        or (sheet == "Server" and number >= 30)
    )


def load_audit_map(path: Path) -> dict[tuple[str, int], list[str]]:
    if not path.is_file():
        return {}
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["AUDIT_PENYESUAIAN"]
    rows = sheet.iter_rows(values_only=True)
    header = [clean(value) for value in next(rows)]
    index = {name: position for position, name in enumerate(header)}
    result: dict[tuple[str, int], list[str]] = defaultdict(list)
    for values in rows:
        sheet_name = clean(values[index["Sheet Sumber"]])
        raw_row = values[index["Baris"]]
        raw_number = values[index["No Inventaris Registrasi"]]
        if not sheet_name or raw_row in (None, ""):
            continue
        number = digits(raw_number)
        if number:
            result[(sheet_name, int(float(raw_row)))].append(number)
    return dict(result)


def load_reference_maps(path: Path) -> tuple[dict[str, float], dict[str, int]]:
    if not path.is_file():
        return {}, {}
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    prices: dict[str, float] = {}
    years: dict[str, int] = {}
    if "AUDIT_HARGA_REFERENSI" in workbook.sheetnames:
        rows = workbook["AUDIT_HARGA_REFERENSI"].iter_rows(values_only=True)
        header = [clean(value) for value in next(rows)]
        index = {name: position for position, name in enumerate(header)}
        for values in rows:
            number = digits(values[index["No Inventaris"]])
            if number:
                prices[number] = float(values[index["Harga Referensi/Unit"]] or 0)
    if "AUDIT_TAHUN_REFERENSI" in workbook.sheetnames:
        rows = workbook["AUDIT_TAHUN_REFERENSI"].iter_rows(values_only=True)
        header = [clean(value) for value in next(rows)]
        index = {name: position for position, name in enumerate(header)}
        for values in rows:
            number = digits(values[index["No Inventaris"]])
            raw_year = values[index["Tahun Referensi"]]
            if number and raw_year not in (None, ""):
                years[number] = int(float(raw_year))
    return prices, years


def load_database_maps(connection) -> dict:
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT kode,nama,harga_referensi FROM "
            "rsns_custom_logistik_non_medis_inventaris_master "
            "WHERE jenis_master='BARANG' AND status='Aktif'"
        )
        masters = cursor.fetchall()
        cursor.execute(
            "SELECT kode,nama FROM rsns_custom_logistik_non_medis_inventaris_master "
            "WHERE jenis_master='UNIT' AND status='Aktif'"
        )
        units = {row["kode"]: row["nama"] for row in cursor.fetchall()}
        cursor.execute(
            "SELECT kode_item,nama_aset,merk_type,kode_unit,nomor_inventaris "
            "FROM rsns_custom_logistik_non_medis_aset WHERE status='Aktif'"
        )
        assets = cursor.fetchall()

    master_by_code = {row["kode"]: row for row in masters}
    for code, details in REQUIRED_ITEM_MASTERS.items():
        master_by_code.setdefault(
            code, {"kode": code, "nama": details["nama"], "harga_referensi": 0}
        )
    master_name: dict[str, set[str]] = defaultdict(set)
    for row in masters:
        master_name[normalized(row["nama"])].add(row["kode"])

    exact_asset: dict[tuple[str, str], set[str]] = defaultdict(set)
    asset_name: dict[str, set[str]] = defaultdict(set)
    for row in assets:
        name_key = normalized(row["nama_aset"])
        brand_key = normalized(row["merk_type"])
        exact_asset[(name_key, brand_key)].add(row["kode_item"])
        asset_name[name_key].add(row["kode_item"])

    return {
        "master_by_code": master_by_code,
        "master_name": master_name,
        "exact_asset": exact_asset,
        "asset_name": asset_name,
        "units": units,
        "assets": assets,
    }


def candidate_code_from_number(number: str) -> str | None:
    raw = digits(number)
    if len(raw) not in {11, 12, 13}:
        return None
    return legacy.item_code_from_inventory(raw)


def resolve_item_code(row: dict, audit_numbers: list[str], maps: dict) -> tuple[str | None, str, list[str]]:
    master = maps["master_by_code"]
    override = ITEM_OVERRIDES.get((row["sheet"], int(row["row"])))
    if override:
        if override not in master:
            return None, "override_master_tidak_ada", [override]
        return override, "override_terverifikasi", [override]
    candidates: list[tuple[str, str]] = []
    for number in audit_numbers:
        code = candidate_code_from_number(number)
        if code in master:
            candidates.append((code, "audit_registrasi"))
    if row.get("inventory"):
        code = candidate_code_from_number(row["inventory"])
        if code in master:
            candidates.append((code, "nomor_list"))

    name_key = normalized(row["name"])
    brand_key = normalized(row["brand"])
    exact_codes = maps["exact_asset"].get((name_key, brand_key), set())
    if len(exact_codes) == 1:
        candidates.append((next(iter(exact_codes)), "nama_merk_database"))
    name_codes = maps["asset_name"].get(name_key, set())
    if len(name_codes) == 1:
        candidates.append((next(iter(name_codes)), "nama_database"))
    master_codes = maps["master_name"].get(name_key, set())
    if len(master_codes) == 1:
        candidates.append((next(iter(master_codes)), "nama_master"))

    # A unique exact name already used in the database is strongest evidence.
    # This also repairs malformed LIST numbers whose embedded item segment is
    # a different, valid master code.
    for preferred_reason in ("nama_merk_database", "nama_database", "nama_master"):
        preferred = {code for code, reason in candidates if reason == preferred_reason}
        if len(preferred) == 1:
            code = next(iter(preferred))
            return code, preferred_reason, sorted({candidate for candidate, _reason in candidates})

    numbered = []
    for code, reason in candidates:
        if reason in {"audit_registrasi", "nomor_list"} and code not in [item[0] for item in numbered]:
            numbered.append((code, reason))
    if len(numbered) == 1:
        return numbered[0][0], numbered[0][1], sorted({code for code, _reason in candidates})
    if not numbered:
        return None, "tidak_ditemukan", []
    return None, "ambigu", [code for code, _reason in numbered]


def parse_and_resolve(path: Path, registration_path: Path, maps: dict) -> tuple[list[dict], list[dict], list[dict]]:
    audit_map = load_audit_map(registration_path)
    parsed = legacy.parse_workbook(path)
    resolved: list[dict] = []
    excluded: list[dict] = []
    issues: list[dict] = []

    for source in parsed:
        if is_historical(source):
            excluded.append(source)
            continue

        row = dict(source)
        row["condition_note"] = ""
        inferred_number = int(row["unit"].replace("XLS-", ""))
        explicit_number = explicit_location_unit(row["location"])
        selected_number = 64 if row["sheet"] == "Selasar Lt.1 Gd lama" else (explicit_number or inferred_number)
        row["unit"] = unit_code(selected_number)
        row["audit_numbers"] = audit_map.get((row["sheet"], int(row["row"])), [])
        code, method, candidates = resolve_item_code(row, row["audit_numbers"], maps)
        row["item_code"] = code
        row["resolution"] = method

        # AC is locked and excluded using either its resolved code or exact name.
        if code in AC_CODES or normalized(row["name"]) == "ac":
            excluded.append(row)
            continue

        if row["unit"] not in maps["units"]:
            issues.append({
                "sheet": row["sheet"], "row": row["row"], "name": row["name"],
                "issue": "unit_tidak_ada", "unit": row["unit"],
            })
        if code is None:
            issues.append({
                "sheet": row["sheet"], "row": row["row"], "name": row["name"],
                "brand": row["brand"], "inventory": row["raw_inventory"],
                "issue": method, "candidates": candidates,
            })
        resolved.append(row)

    return resolved, excluded, issues


def assign_numbers(rows: list[dict], maps: dict, price_map: dict[str, float], year_map: dict[str, int]) -> None:
    used: dict[tuple[str, str], set[int]] = defaultdict(set)
    for asset in maps["assets"]:
        if asset["kode_unit"] == PRESERVED_UNIT or asset["kode_item"] in AC_CODES:
            number = digits(asset["nomor_inventaris"])
            if len(number) >= 3:
                used[(asset["kode_unit"], asset["kode_item"])].add(int(number[-3:]))

    for row in rows:
        key = (row["unit"], row["item_code"])
        preferred_number = row["audit_numbers"][0] if row["audit_numbers"] else (row["inventory"] or "")
        requested = source_sequence(preferred_number)
        assigned: list[int] = []
        for offset in range(int(row["quantity"])):
            candidate = requested + offset
            while candidate in used[key] and candidate <= 999:
                candidate += 1
            if candidate > 999:
                candidate = 1
                while candidate in used[key] and candidate <= 999:
                    candidate += 1
            if candidate > 999:
                raise ValueError(f"Nomor urut habis untuk unit {row['unit']} item {row['item_code']}")
            used[key].add(candidate)
            assigned.append(candidate)

        row["numbers"] = [row["unit"] + "2" + row["item_code"] + f"{seq:03d}" for seq in assigned]
        quantity = int(row["quantity"])
        purchase = float(row["price"] or 0) / quantity
        reference_number = preferred_number
        master_reference = float(maps["master_by_code"][row["item_code"]]["harga_referensi"] or 0)
        row["purchase_price"] = purchase
        row["reference_price"] = 0.0 if purchase > 0 else float(price_map.get(reference_number, master_reference) or 0)
        row["display_price"] = purchase or row["reference_price"]
        row["year_reference"] = 0
        if row["year"] is None and reference_number in year_map:
            row["year"] = year_map[reference_number]
            row["year_reference"] = 1


def sql_value(value) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, (int, float)):
        return str(value)
    escaped = str(value).replace("\\", "\\\\").replace("'", "''")
    return f"'{escaped}'"


def build_sql(rows: list[dict], commit: bool = True) -> str:
    statements = [
        "SET NAMES utf8mb4;",
        "START TRANSACTION;",
    ]
    for code, details in REQUIRED_ITEM_MASTERS.items():
        values = [
            "BARANG", code, None, "2", details["nama"], details["kelompok"], details["jenis"],
            details["barang"], details["nama_kelompok"], details["nama_jenis"], "B", "Aktif",
        ]
        statements.append(
            "INSERT INTO rsns_custom_logistik_non_medis_inventaris_master "
            "(jenis_master,kode,kode_inventaris,kode_kategori,nama,kode_kelompok,kode_jenis,kode_barang,"
            "nama_kelompok,nama_jenis,kib_jenis,status,tgl_input) VALUES ("
            + ",".join(sql_value(value) for value in values) + ",NOW()) "
            "ON DUPLICATE KEY UPDATE nama=VALUES(nama),kode_kelompok=VALUES(kode_kelompok),"
            "kode_jenis=VALUES(kode_jenis),kode_barang=VALUES(kode_barang),nama_kelompok=VALUES(nama_kelompok),"
            "nama_jenis=VALUES(nama_jenis),status='Aktif';"
        )
    statements.extend([
        "DELETE FROM rsns_custom_logistik_non_medis_aset "
        "WHERE kode_unit<>'04' AND kode_item NOT IN ('021300','021301');",
        "DELETE FROM rsns_custom_logistik_non_medis_asset_groups "
        "WHERE kode_unit<>'04' AND kode_item NOT IN ('021300','021301');",
    ])
    for index, row in enumerate(rows, 1):
        group_code = f"GRP-NONAC-20260825-{index:04d}"
        group_values = [
            group_code, row["item_code"], "2", row["name"], row["unit"], row["location"],
            None, row["year"], row["source"], "Unit", row["display_price"], len(row["numbers"]),
            row["numbers"][0], row["numbers"][-1], "Aktif", SYNC_USER,
        ]
        statements.append(
            "INSERT INTO rsns_custom_logistik_non_medis_asset_groups "
            "(kode_group,kode_item,kode_kategori,nama_group,kode_unit,lokasi_fisik,tanggal_perolehan,tahun_beli,"
            "sumber_perolehan,satuan,harga_satuan,jumlah,nomor_awal,nomor_akhir,status,tgl_input,user_input) VALUES ("
            + ",".join(sql_value(value) for value in group_values[:-1]) + ",NOW(),"
            + sql_value(group_values[-1]) + ");"
        )
        statements.append("SET @nonac_group_id = LAST_INSERT_ID();")
        for number in row["numbers"]:
            notes = [f"Sumber LIST: {row['sheet']} baris {row['row']}"]
            if row.get("raw_inventory"):
                notes.append(f"Nomor LIST: {row['raw_inventory']}")
            if row["unit"] == "102":
                notes.append("Kode 102 perlu verifikasi: Multazam 4 atau Multazam 12")
            if row["reference_price"]:
                notes.append("Harga merupakan referensi, bukan bukti harga beli")
            if row["year_reference"]:
                notes.append("Tahun merupakan referensi")
            asset_values = [
                "AST-" + number, number, row.get("inventory"), row["item_code"], "2", row["name"],
                row["brand"] or None, None, row["year"], row["year_reference"], "Unit", 1,
                row["purchase_price"], row["reference_price"],
                "Referensi audit/master barang" if row["reference_price"] else None,
                row["source"], row["unit"], row["location"], row["condition"], " | ".join(notes),
                "Aktif", row["purchase_price"], SYNC_USER, "B", row["brand"] or None,
            ]
            statements.append(
                "INSERT INTO rsns_custom_logistik_non_medis_aset "
                "(asset_group_id,kode_aset,nomor_inventaris,nomor_dokumen,kode_item,kode_kategori_aset,nama_aset,"
                "merk_type,tanggal_perolehan,tahun_beli,tahun_beli_referensi,satuan,jumlah,harga_beli,"
                "harga_referensi_import,harga_referensi_keterangan,sumber_perolehan,kode_unit,lokasi_fisik,"
                "status_kondisi,keterangan_inventaris,status,nilai_buku,tgl_input,user_input,kib_jenis,kib_merk) VALUES "
                "(@nonac_group_id," + ",".join(sql_value(value) for value in asset_values[:-3]) + ",NOW(),"
                + ",".join(sql_value(value) for value in asset_values[-3:]) + ");"
            )
    statements.extend([
        "COMMIT;" if commit else "ROLLBACK;",
        "SELECT COUNT(*) AS groups_non_ac,COALESCE(SUM(jumlah),0) AS qty_non_ac "
        "FROM rsns_custom_logistik_non_medis_asset_groups WHERE status='Aktif' AND kode_item NOT IN ('021300','021301');",
        "SELECT COUNT(*) AS assets_all FROM rsns_custom_logistik_non_medis_aset WHERE status='Aktif';",
    ])
    return "\n".join(statements) + "\n"


def validate_database(rows: list[dict], connection) -> dict:
    numbers = [number for row in rows for number in row["numbers"]]
    if len(numbers) != len(set(numbers)):
        raise ValueError("Nomor inventaris hasil non-AC tidak unik")
    with connection.cursor() as cursor:
        histories = {}
        for table in [
            "rsns_custom_logistik_non_medis_aset_mutasi",
            "rsns_custom_logistik_non_medis_aset_pemeliharaan",
            "rsns_custom_logistik_non_medis_aset_penghapusan",
            "rsns_custom_logistik_non_medis_aset_penyusutan",
            "rsns_custom_logistik_non_medis_aset_sensus",
        ]:
            cursor.execute(
                f"SELECT COUNT(*) total FROM `{table}` h JOIN rsns_custom_logistik_non_medis_aset a "
                "ON a.kode_aset=h.kode_aset WHERE a.kode_unit<>'04' "
                "AND a.kode_item NOT IN ('021300','021301')"
            )
            histories[table] = int(cursor.fetchone()["total"])
        if any(histories.values()):
            raise RuntimeError(f"Barang target memiliki riwayat transaksi: {histories}")

        placeholders = ",".join(["%s"] * len(numbers))
        cursor.execute(
            f"SELECT nomor_inventaris,kode_item,kode_unit FROM rsns_custom_logistik_non_medis_aset "
            f"WHERE nomor_inventaris IN ({placeholders}) "
            "AND (kode_unit='04' OR kode_item IN ('021300','021301'))",
            numbers,
        )
        conflicts = cursor.fetchall()
        if conflicts:
            raise RuntimeError(f"Nomor bertabrakan dengan aset yang dipertahankan: {conflicts[:10]}")
    return {"history": histories, "numbers": len(numbers)}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--registration-workbook", type=Path, default=DEFAULT_REGISTRATION)
    parser.add_argument("--reference-workbook", type=Path, default=DEFAULT_REFERENCE)
    parser.add_argument("--host", default="localhost")
    parser.add_argument("--port", type=int, default=3306)
    parser.add_argument("--user", default="root")
    parser.add_argument("--password", default="")
    parser.add_argument("--database", default="mlite_rsns")
    parser.add_argument("--sql-output", type=Path)
    parser.add_argument("--rollback-output", type=Path)
    parser.add_argument("--report", type=Path)
    args = parser.parse_args()

    connection = pymysql.connect(
        host=args.host, port=args.port, user=args.user, password=args.password,
        database=args.database, charset="utf8mb4", autocommit=False,
        cursorclass=pymysql.cursors.DictCursor,
    )
    try:
        maps = load_database_maps(connection)
        rows, excluded, issues = parse_and_resolve(args.workbook, args.registration_workbook, maps)
        price_map, year_map = load_reference_maps(args.reference_workbook)
        if not issues:
            assign_numbers(rows, maps, price_map, year_map)
            validation = validate_database(rows, connection)
        else:
            validation = {"history": {}, "numbers": 0}
    finally:
        connection.rollback()
        connection.close()

    methods: dict[str, int] = defaultdict(int)
    for row in rows:
        methods[row["resolution"]] += 1
    report = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_file": str(args.workbook),
        "target_groups": len(rows),
        "target_quantity": sum(int(row["quantity"]) for row in rows),
        "excluded_groups": len(excluded),
        "excluded_quantity": sum(int(row.get("quantity", 1)) for row in excluded),
        "missing_inventory_groups": sum(not row.get("inventory") for row in rows),
        "missing_inventory_quantity": sum(int(row["quantity"]) for row in rows if not row.get("inventory")),
        "resolution_methods": dict(sorted(methods.items())),
        "issues": issues,
        "validation": validation,
    }
    if not issues:
        sql = build_sql(rows)
        report["target_number_hash"] = hashlib.md5("\n".join(
            number for row in rows for number in row["numbers"]
        ).encode()).hexdigest()
        if args.sql_output:
            args.sql_output.parent.mkdir(parents=True, exist_ok=True)
            args.sql_output.write_text(sql, encoding="utf-8")
        if args.rollback_output:
            args.rollback_output.parent.mkdir(parents=True, exist_ok=True)
            args.rollback_output.write_text(build_sql(rows, commit=False), encoding="utf-8")
    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 2 if issues else 0


if __name__ == "__main__":
    raise SystemExit(main())
