"""Reconcile grouped quantities from the per-unit workbook into individual assets.

Dry-run is the default. Pass --apply to commit changes to the local mlite_rsns DB.
Only workbook rows with Jumlah > 1 are considered.
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from copy import deepcopy
from pathlib import Path

import pymysql
from openpyxl import load_workbook


DEFAULT_ZIP_DIR = Path(
    r"C:\Users\ASUS\AppData\Local\Temp\registrasi_aset_per_unit_20260823_1340"
)


def as_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def load_multi_quantity_rows(root: Path) -> list[dict]:
    result: list[dict] = []
    for workbook_path in sorted(root.glob("*.xlsx")):
        if workbook_path.name == "Daftar_Registrasi_Aset_Per_Unit_Final.xlsx":
            continue
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        sheet = workbook["IMPORT_REGISTRASI_ASET"]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows)
        index = {str(value).strip(): idx for idx, value in enumerate(header) if value is not None}
        for row in rows:
            if not any(value is not None for value in row):
                continue
            quantity = int(row[index["Jumlah"]] or 0)
            if quantity <= 1:
                continue
            result.append(
                {
                    "file": workbook_path.name,
                    "source_inventory": as_text(row[index["No. Inventaris"]]),
                    "unit": as_text(row[index["Kode Unit"]]),
                    "name": as_text(row[index["Nama Aset"]]),
                    "quantity": quantity,
                }
            )
    return result


def fetch_all(cursor, sql: str, params=()) -> list[dict]:
    cursor.execute(sql, params)
    return list(cursor.fetchall())


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--workbooks", type=Path, default=DEFAULT_ZIP_DIR)
    args = parser.parse_args()

    source_rows = load_multi_quantity_rows(args.workbooks)
    if not source_rows:
        raise RuntimeError("Tidak ditemukan baris dengan Jumlah > 1.")

    connection = pymysql.connect(
        host="localhost",
        user="root",
        password="",
        database="mlite_rsns",
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=False,
    )

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT COUNT(*) AS rows_count, COALESCE(SUM(jumlah),0) AS physical "
                "FROM rsns_custom_logistik_non_medis_aset WHERE status='Aktif'"
            )
            before = cursor.fetchone()

            source_by_item: dict[tuple[str, str], list[dict]] = defaultdict(list)
            for row in source_rows:
                source_by_item[(row["unit"], row["name"])].append(row)

            group_targets: dict[int, dict] = {}
            affected_item_codes: set[tuple[str, str]] = set()

            for (unit, name), rows in sorted(source_by_item.items()):
                existing = fetch_all(
                    cursor,
                    "SELECT * FROM rsns_custom_logistik_non_medis_aset "
                    "WHERE status='Aktif' AND kode_unit=%s AND nama_aset=%s "
                    "ORDER BY nomor_inventaris, id",
                    (unit, name),
                )
                if len(existing) != len(rows):
                    raise RuntimeError(
                        f"Anchor {unit}/{name} tidak cocok: workbook={len(rows)}, DB={len(existing)}"
                    )

                for source, asset in zip(rows, existing):
                    group_id = int(asset["asset_group_id"] or 0)
                    if not group_id:
                        raise RuntimeError(f"Aset {asset['id']} belum memiliki asset_group_id.")
                    target = group_targets.setdefault(
                        group_id,
                        {
                            "unit": unit,
                            "name": name,
                            "kode_item": asset["kode_item"],
                            "quantity": 0,
                            "source_rows": [],
                        },
                    )
                    target["quantity"] += int(source["quantity"])
                    target["source_rows"].append(source)
                    affected_item_codes.add((unit, asset["kode_item"]))

            affected_group_ids = sorted(group_targets)
            placeholders = ",".join(["%s"] * len(affected_group_ids))
            existing_assets = fetch_all(
                cursor,
                f"SELECT * FROM rsns_custom_logistik_non_medis_aset "
                f"WHERE status='Aktif' AND asset_group_id IN ({placeholders}) "
                "ORDER BY asset_group_id, nomor_inventaris, id",
                affected_group_ids,
            )
            assets_by_group: dict[int, list[dict]] = defaultdict(list)
            for asset in existing_assets:
                assets_by_group[int(asset["asset_group_id"])].append(asset)

            for unit, item_code in affected_item_codes:
                rows = fetch_all(
                    cursor,
                    "SELECT id,asset_group_id,nama_aset FROM rsns_custom_logistik_non_medis_aset "
                    "WHERE status='Aktif' AND kode_unit=%s AND kode_item=%s",
                    (unit, item_code),
                )
                unexpected = [
                    row for row in rows if int(row["asset_group_id"] or 0) not in group_targets
                ]
                if unexpected:
                    raise RuntimeError(
                        f"Kode {unit}/{item_code} juga dipakai grup lain; koreksi dihentikan."
                    )

            assignments: dict[int, list[str]] = {}
            for unit, item_code in sorted(affected_item_codes):
                item_groups = [
                    group_id
                    for group_id, target in group_targets.items()
                    if target["unit"] == unit and target["kode_item"] == item_code
                ]
                item_groups.sort(
                    key=lambda group_id: min(
                        asset["nomor_inventaris"] for asset in assets_by_group[group_id]
                    )
                )
                sample_number = min(
                    asset["nomor_inventaris"]
                    for group_id in item_groups
                    for asset in assets_by_group[group_id]
                )
                if not sample_number or len(sample_number) < 4:
                    raise RuntimeError(f"Nomor inventaris {unit}/{item_code} tidak valid.")
                prefix = sample_number[:-3]
                sequence = 1
                for group_id in item_groups:
                    quantity = group_targets[group_id]["quantity"]
                    assignments[group_id] = [
                        f"{prefix}{number:03d}"
                        for number in range(sequence, sequence + quantity)
                    ]
                    sequence += quantity

            planned_new = sum(
                group_targets[group_id]["quantity"] - len(assets_by_group[group_id])
                for group_id in affected_group_ids
            )

            print(
                f"BEFORE rows={before['rows_count']} physical={before['physical']} "
                f"affected_groups={len(affected_group_ids)} planned_new={planned_new}"
            )
            for group_id in affected_group_ids:
                target = group_targets[group_id]
                numbers = assignments[group_id]
                print(
                    f"GROUP {group_id} | {target['unit']} | {target['name']} | "
                    f"{len(assets_by_group[group_id])} -> {target['quantity']} | "
                    f"{numbers[0]} - {numbers[-1]}"
                )

            if not args.apply:
                connection.rollback()
                print("DRY_RUN_OK")
                return 0

            # Move existing unique keys out of the way before assigning new ranges.
            for asset in existing_assets:
                cursor.execute(
                    "UPDATE rsns_custom_logistik_non_medis_aset "
                    "SET nomor_inventaris=%s,kode_aset=%s WHERE id=%s",
                    (f"TMP-{asset['id']}", f"TMP-AST-{asset['id']}", asset["id"]),
                )

            insert_columns = [
                column
                for column in existing_assets[0].keys()
                if column != "id"
            ]
            insert_sql = (
                "INSERT INTO rsns_custom_logistik_non_medis_aset (`"
                + "`,`".join(insert_columns)
                + "`) VALUES ("
                + ",".join(["%s"] * len(insert_columns))
                + ")"
            )

            for group_id in affected_group_ids:
                current_assets = assets_by_group[group_id]
                target_quantity = group_targets[group_id]["quantity"]
                numbers = assignments[group_id]
                total_purchase = sum(float(asset["harga_beli"] or 0) for asset in current_assets)
                total_reference = sum(
                    float(asset["harga_referensi_import"] or 0) for asset in current_assets
                )
                purchase_per_unit = total_purchase / target_quantity if total_purchase else 0
                reference_per_unit = total_reference / target_quantity if total_reference else 0
                displayed_unit_price = purchase_per_unit or reference_per_unit
                template = deepcopy(current_assets[0])

                for index, number in enumerate(numbers):
                    values = deepcopy(current_assets[index] if index < len(current_assets) else template)
                    values["nomor_inventaris"] = number
                    values["kode_aset"] = f"AST-{number}"
                    values["jumlah"] = 1
                    values["harga_beli"] = purchase_per_unit
                    values["harga_referensi_import"] = reference_per_unit
                    values["user_input"] = "reconcile_per_unit_20260823"
                    if index < len(current_assets):
                        cursor.execute(
                            "UPDATE rsns_custom_logistik_non_medis_aset SET "
                            "nomor_inventaris=%s,kode_aset=%s,jumlah=1,harga_beli=%s," 
                            "harga_referensi_import=%s,user_input=%s WHERE id=%s",
                            (
                                values["nomor_inventaris"],
                                values["kode_aset"],
                                values["harga_beli"],
                                values["harga_referensi_import"],
                                values["user_input"],
                                current_assets[index]["id"],
                            ),
                        )
                    else:
                        cursor.execute(
                            insert_sql,
                            [values[column] for column in insert_columns],
                        )

                cursor.execute(
                    "UPDATE rsns_custom_logistik_non_medis_asset_groups "
                    "SET jumlah=%s,nomor_awal=%s,nomor_akhir=%s,harga_satuan=%s," 
                    "user_input='reconcile_per_unit_20260823' WHERE id=%s",
                    (
                        target_quantity,
                        numbers[0],
                        numbers[-1],
                        displayed_unit_price,
                        group_id,
                    ),
                )

            cursor.execute(
                "SELECT COUNT(*) AS rows_count, COALESCE(SUM(jumlah),0) AS physical, "
                "COUNT(DISTINCT nomor_inventaris) AS unique_numbers, "
                "SUM(harga_beli) AS total_purchase, "
                "SUM(COALESCE(NULLIF(harga_beli,0),harga_referensi_import)) AS total_display "
                "FROM rsns_custom_logistik_non_medis_aset WHERE status='Aktif'"
            )
            after = cursor.fetchone()
            if int(after["rows_count"]) != 6335 or int(after["physical"]) != 6335:
                raise RuntimeError(f"Total akhir tidak sesuai: {after}")
            if int(after["unique_numbers"]) != int(after["rows_count"]):
                raise RuntimeError("Nomor inventaris akhir tidak unik.")

            connection.commit()
            print(
                f"APPLIED rows={after['rows_count']} physical={after['physical']} "
                f"unique_numbers={after['unique_numbers']} "
                f"total_purchase={after['total_purchase']} total_display={after['total_display']}"
            )
            return 0
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


if __name__ == "__main__":
    raise SystemExit(main())
