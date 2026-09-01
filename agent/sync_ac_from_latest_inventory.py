#!/usr/bin/env python3
"""Synchronize only AC inventory from the latest LIST workbook.

The workbook is data input only. Historical snapshots in the Radiologi and
Logistik sheets are excluded explicitly; all non-AC database rows are kept.
"""

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
import pymysql

import import_inventaris_non_medis_20260813 as legacy


DEFAULT_WORKBOOK = Path(r"C:\Users\ASUS\Downloads\LIST INVENTARIS_NON MEDIS.xlsx")
DEFAULT_REFERENCE_WORKBOOK = Path(
    r"C:\Users\ASUS\Downloads\FINAL_Import_Registrasi_Aset_Dengan_Referensi.xlsx"
)

# These are older snapshots of the same rooms. The latest rows remain:
# Radiologi April 2026 and Logistik April 2026.
HISTORICAL_AC_ROWS = {
    ("radiologi", 10),
    ("radiologi", 49),
    ("logistik", 116),
    ("logistik", 190),
}

# User-confirmed newest additions. Unit 108 already has one AC in the workbook,
# so six rows are added. The purchase details below preserve the data entered
# on the VPS while this reconciliation was prepared.
REQUIRED_ROOM_AC = {
    "61": "Marwa 19",
    "62": "Marwa 20",
    "63": "Marwa 21",
    "106": "Multazam 16",
    "107": "Multazam 17",
    "108": "Multazam 18",
    "109": "Multazam 19",
}
ADDITION_UNIT_CODES = ("61", "62", "63", "106", "107", "109")
ADDITION_NAME = "AC split TCL 1 PK"
ADDITION_GROUP_NAME = "AC split"
ADDITION_ITEM_CODE = "021301"
ADDITION_PRICE = 3_250_000.0
ADDITION_DATE = "2026-07-27"
ADDITION_YEAR = 2026

EXPECTED_SOURCE_GROUPS = 105
EXPECTED_SOURCE_QUANTITY = 118
SYNC_USER = "sync_ac_list_20260825"


def clean_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def digits(value) -> str:
    return re.sub(r"\D", "", clean_text(value))


def unit_code(number: int) -> str:
    return f"{number:02d}" if number < 100 else str(number)


def explicit_location_unit(location: str) -> int | None:
    # Only parenthesized codes are considered authoritative. Plain room
    # numbers such as "Farmasi Rawat Jalan 2" are not unit codes.
    matches = re.findall(r"\(\s*'?\s*(\d{1,3})\s*\)", location)
    if not matches:
        return None
    candidate = int(matches[-1])
    return candidate if 1 <= candidate <= 142 else None


def source_sequence(inventory: str) -> int:
    value = digits(inventory)
    if len(value) < 2:
        return 1
    result = int(value[-2:])
    return result if result > 0 else 1


def parse_asuransi_ac(path: Path) -> list[dict]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Asuransi"]
    headers = None
    result = []
    for row_number, values_tuple in enumerate(sheet.iter_rows(values_only=True), 1):
        values = list(values_tuple)
        normalized = [clean_text(value).upper() for value in values]
        discovered = legacy.header_map(normalized)
        if discovered:
            headers = discovered
            continue
        if not headers:
            continue

        def get(key):
            index = headers.get(key, -1)
            return values[index] if 0 <= index < len(values) else None

        sequence = get("no")
        name = clean_text(get("name"))
        if not isinstance(sequence, (int, float)) or sequence != int(sequence):
            continue
        if name.casefold() != "ac":
            continue
        raw_inventory = clean_text(get("inventory"))
        inventory = legacy.inventory_digits(raw_inventory)
        quantity = max(1, int(round(legacy.number(get("quantity"), 1))))
        raw_year = int(legacy.number(get("year"), 0))
        year = raw_year if 1900 <= raw_year <= datetime.now().year else None
        condition, condition_note = legacy.normalize_condition(get("condition"))
        result.append(
            {
                "sheet": sheet.title,
                "row": row_number,
                "inventory": inventory or None,
                "raw_inventory": raw_inventory,
                "name": "AC",
                "brand": clean_text(get("brand"))[:150],
                "quantity": quantity,
                "price": max(0.0, legacy.number(get("price"), 0)),
                "year": year,
                "source": legacy.normalize_source(get("source")),
                "condition": condition,
                "condition_note": condition_note,
                "unit": "04",
                "location": "Asuransi",
                "kind": "workbook",
            }
        )
    return result


def parse_source(path: Path) -> tuple[list[dict], list[dict]]:
    parsed = legacy.parse_workbook(path)
    ac_rows = []
    excluded = []
    for row in parsed:
        if clean_text(row["name"]).casefold() != "ac":
            continue
        if (row["sheet"], row["row"]) in HISTORICAL_AC_ROWS:
            excluded.append(row)
            continue

        inferred_number = int(row["unit"].replace("XLS-", ""))
        location_number = explicit_location_unit(row["location"])
        selected_number = location_number or inferred_number
        normalized = dict(row)
        normalized["unit"] = unit_code(selected_number)
        normalized["condition_note"] = ""
        normalized["kind"] = "workbook"
        ac_rows.append(normalized)

    ac_rows.extend(parse_asuransi_ac(path))
    if len(ac_rows) != EXPECTED_SOURCE_GROUPS:
        raise ValueError(
            f"Kelompok AC sumber={len(ac_rows)}, seharusnya {EXPECTED_SOURCE_GROUPS}."
        )
    source_quantity = sum(int(row["quantity"]) for row in ac_rows)
    if source_quantity != EXPECTED_SOURCE_QUANTITY:
        raise ValueError(
            f"Jumlah AC sumber={source_quantity}, seharusnya {EXPECTED_SOURCE_QUANTITY}."
        )
    return ac_rows, excluded


def load_reference_maps(path: Path) -> tuple[dict[str, float], dict[str, int]]:
    if not path.is_file():
        return {}, {}
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)

    price_map = {}
    rows = workbook["AUDIT_HARGA_REFERENSI"].iter_rows(values_only=True)
    headers = [clean_text(value) for value in next(rows)]
    index = {value: position for position, value in enumerate(headers)}
    for row in rows:
        number = digits(row[index["No Inventaris"]])
        if number:
            price_map[number] = float(row[index["Harga Referensi/Unit"]] or 0)

    year_map = {}
    rows = workbook["AUDIT_TAHUN_REFERENSI"].iter_rows(values_only=True)
    headers = [clean_text(value) for value in next(rows)]
    index = {value: position for position, value in enumerate(headers)}
    for row in rows:
        number = digits(row[index["No Inventaris"]])
        year = row[index["Tahun Referensi"]]
        if number and year not in (None, ""):
            year_map[number] = int(float(year))
    return price_map, year_map


def item_code_for(row: dict) -> str:
    if row.get("item_code_override"):
        return row["item_code_override"]
    inventory = row.get("inventory") or ""
    return legacy.item_code_from_inventory(inventory) if inventory else "021300"


def enrich_rows(rows: list[dict], connection, price_map: dict[str, float], year_map: dict[str, int]) -> list[dict]:
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT kode,harga_referensi FROM rsns_custom_logistik_non_medis_inventaris_master "
            "WHERE jenis_master='BARANG' AND kode IN ('021300','021301')"
        )
        master_prices = {row["kode"]: float(row["harga_referensi"] or 0) for row in cursor.fetchall()}
        cursor.execute(
            "SELECT kode,nama FROM rsns_custom_logistik_non_medis_inventaris_master "
            "WHERE jenis_master='UNIT' AND status='Aktif'"
        )
        units = {row["kode"]: row["nama"] for row in cursor.fetchall()}
        cursor.execute(
            "SELECT kode_unit,kode_item,nomor_inventaris "
            "FROM rsns_custom_logistik_non_medis_aset "
            "WHERE status='Aktif' AND kode_item IN ('021300','021301') "
            "AND TRIM(LOWER(nama_aset))<>'ac' "
            "AND NOT (kode_item='021301' AND TRIM(nama_aset) LIKE 'AC split TCL 1 PK%%' "
            "AND kode_unit IN ('61','62','63','106','107','109'))"
        )
        preserved_ac_variants = cursor.fetchall()

    missing_units = sorted({row["unit"] for row in rows} - set(units))
    if missing_units:
        raise ValueError(f"Master unit tidak ditemukan: {', '.join(missing_units)}")

    existing_required = {row["unit"] for row in rows if row["unit"] in REQUIRED_ROOM_AC}
    for code, name in REQUIRED_ROOM_AC.items():
        if code in existing_required:
            continue
        rows.append(
            {
                "sheet": "Tambahan pengguna",
                "row": 0,
                "inventory": None,
                "raw_inventory": "",
                "name": ADDITION_NAME,
                "group_name": ADDITION_GROUP_NAME,
                "item_code_override": ADDITION_ITEM_CODE,
                "brand": "",
                "quantity": 1,
                "price": ADDITION_PRICE,
                "year": ADDITION_YEAR,
                "acquisition_date": ADDITION_DATE,
                "unit_label": "Buah",
                "pic": "x",
                "source": "Beli",
                "condition": "Baik",
                "condition_note": "",
                "unit": code,
                "location": name,
                "kind": "user_addition",
            }
        )

    used_sequences: dict[tuple[str, str], set[int]] = defaultdict(set)
    for asset in preserved_ac_variants:
        number = digits(asset["nomor_inventaris"])
        if len(number) >= 3:
            used_sequences[(asset["kode_unit"], asset["kode_item"])].add(int(number[-3:]))
    enriched = []
    for row in rows:
        item_code = item_code_for(row)
        if item_code not in master_prices:
            raise ValueError(f"Master AC {item_code} tidak ditemukan.")
        key = (row["unit"], item_code)
        requested = source_sequence(row.get("inventory") or "")
        assigned = []
        for offset in range(int(row["quantity"])):
            candidate = requested + offset
            while candidate in used_sequences[key]:
                candidate += 1
            if candidate > 999:
                raise ValueError(f"Nomor urut AC habis untuk unit {row['unit']}.")
            used_sequences[key].add(candidate)
            assigned.append(candidate)

        raw_number = row.get("inventory") or ""
        raw_unit_price = float(row.get("price") or 0) / int(row["quantity"])
        if raw_unit_price > 0:
            purchase_price = raw_unit_price
            reference_price = 0.0
        else:
            purchase_price = 0.0
            reference_price = price_map.get(raw_number, master_prices.get(item_code, 0.0))

        year = row.get("year")
        year_reference = 0
        if year is None and raw_number in year_map:
            year = year_map[raw_number]
            year_reference = 1

        normalized_numbers = [
            row["unit"] + "2" + item_code + f"{sequence:03d}" for sequence in assigned
        ]
        updated = dict(row)
        updated.update(
            {
                "item_code": item_code,
                "unit_name": units[row["unit"]],
                "numbers": normalized_numbers,
                "purchase_price": purchase_price,
                "reference_price": reference_price,
                "display_price": purchase_price or reference_price,
                "year": year,
                "year_reference": year_reference,
            }
        )
        enriched.append(updated)

    return enriched


def sql_value(value) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, (int, float)):
        return str(value)
    escaped = str(value).replace("\\", "\\\\").replace("'", "''")
    return f"'{escaped}'"


def build_sql(rows: list[dict]) -> str:
    statements = [
        "SET NAMES utf8mb4;",
        "START TRANSACTION;",
        "DELETE FROM rsns_custom_logistik_non_medis_aset WHERE TRIM(LOWER(nama_aset))='ac' "
        "OR (kode_item='021301' AND TRIM(nama_aset) LIKE 'AC split TCL 1 PK%' "
        "AND kode_unit IN ('61','62','63','106','107','109'));",
        "DELETE FROM rsns_custom_logistik_non_medis_asset_groups WHERE TRIM(LOWER(nama_group))='ac' "
        "OR (kode_item='021301' AND TRIM(LOWER(nama_group))='ac split' "
        "AND kode_unit IN ('61','62','63','106','107','109'));",
    ]
    for index, row in enumerate(rows, 1):
        group_code = f"GRP-AC-20260825-{index:03d}"
        group_values = [
            group_code,
            row["item_code"],
            "2",
            row.get("group_name", "AC"),
            row["unit"],
            row["location"],
            row.get("acquisition_date"),
            row["year"],
            row["source"],
            row.get("unit_label", "Unit"),
            row["display_price"],
            len(row["numbers"]),
            row["numbers"][0],
            row["numbers"][-1],
            "Aktif",
            SYNC_USER,
        ]
        statements.append(
            "INSERT INTO rsns_custom_logistik_non_medis_asset_groups "
            "(kode_group,kode_item,kode_kategori,nama_group,kode_unit,lokasi_fisik,tanggal_perolehan,tahun_beli,"
            "sumber_perolehan,satuan,harga_satuan,jumlah,nomor_awal,nomor_akhir,status,tgl_input,user_input) VALUES ("
            + ",".join(sql_value(value) for value in group_values[:-1])
            + ",NOW(),"
            + sql_value(group_values[-1])
            + ");"
        )
        statements.append("SET @ac_group_id = LAST_INSERT_ID();")

        for asset_index, number in enumerate(row["numbers"]):
            source_note = (
                f"Sumber LIST: {row['sheet']} baris {row['row']}"
                if row["kind"] == "workbook"
                else "Penambahan pengguna: satu AC per ruang"
            )
            notes = [source_note]
            if row.get("raw_inventory"):
                notes.append(f"Nomor LIST: {row['raw_inventory']}")
            if row["unit"] == "102":
                notes.append("Kode 102 perlu verifikasi: Multazam 4 atau Multazam 12")
            if row["reference_price"]:
                notes.append("Harga merupakan referensi, bukan bukti harga beli")
            if row["year_reference"]:
                notes.append("Tahun merupakan referensi")

            source_document = row.get("inventory") or None
            asset_values = [
                "AST-" + number,
                number,
                source_document,
                row["item_code"],
                "2",
                row["name"],
                row["brand"] or None,
                row.get("acquisition_date"),
                row["year"],
                row["year_reference"],
                row.get("unit_label", "Unit"),
                1,
                row["purchase_price"],
                row["reference_price"],
                "Referensi dari audit/master AC" if row["reference_price"] else None,
                row["source"],
                row["unit"],
                row["location"],
                row.get("pic"),
                row["condition"],
                " | ".join(notes),
                "Aktif",
                row["purchase_price"],
                SYNC_USER,
                "B",
                row["brand"] or None,
            ]
            statements.append(
                "INSERT INTO rsns_custom_logistik_non_medis_aset "
                "(asset_group_id,kode_aset,nomor_inventaris,nomor_dokumen,kode_item,kode_kategori_aset,nama_aset,"
                "merk_type,tanggal_perolehan,tahun_beli,tahun_beli_referensi,satuan,jumlah,harga_beli,harga_referensi_import,"
                "harga_referensi_keterangan,sumber_perolehan,kode_unit,lokasi_fisik,pic,status_kondisi,"
                "keterangan_inventaris,status,nilai_buku,tgl_input,user_input,kib_jenis,kib_merk) VALUES "
                "(@ac_group_id,"
                + ",".join(sql_value(value) for value in asset_values[:-3])
                + ",NOW(),"
                + ",".join(sql_value(value) for value in asset_values[-3:])
                + ");"
            )
    statements.extend(
        [
            "COMMIT;",
            "SELECT COUNT(*) AS groups_ac,COALESCE(SUM(jumlah),0) AS group_quantity "
            "FROM rsns_custom_logistik_non_medis_asset_groups WHERE status='Aktif' AND TRIM(LOWER(nama_group))='ac';",
            "SELECT COUNT(*) AS assets_ac,COALESCE(SUM(jumlah),0) AS asset_quantity "
            "FROM rsns_custom_logistik_non_medis_aset WHERE status='Aktif' AND TRIM(LOWER(nama_aset))='ac';",
            "SELECT COUNT(*) AS groups_all_ac,COALESCE(SUM(jumlah),0) AS all_ac_quantity "
            "FROM rsns_custom_logistik_non_medis_asset_groups WHERE status='Aktif' AND kode_item IN ('021300','021301');",
        ]
    )
    return "\n".join(statements) + "\n"


def validate_database(rows: list[dict], connection) -> dict:
    numbers = [number for row in rows for number in row["numbers"]]
    if len(numbers) != len(set(numbers)):
        raise ValueError("Nomor inventaris hasil sinkronisasi tidak unik.")

    with connection.cursor() as cursor:
        history_counts = {}
        for table in [
            "rsns_custom_logistik_non_medis_aset_mutasi",
            "rsns_custom_logistik_non_medis_aset_pemeliharaan",
            "rsns_custom_logistik_non_medis_aset_penghapusan",
            "rsns_custom_logistik_non_medis_aset_penyusutan",
            "rsns_custom_logistik_non_medis_aset_sensus",
        ]:
            cursor.execute(
                f"SELECT COUNT(*) total FROM `{table}` h "
                "JOIN rsns_custom_logistik_non_medis_aset a ON a.kode_aset=h.kode_aset "
                "WHERE TRIM(LOWER(a.nama_aset))='ac' OR "
                "(a.kode_item='021301' AND TRIM(a.nama_aset) LIKE 'AC split TCL 1 PK%' "
                "AND a.kode_unit IN ('61','62','63','106','107','109'))"
            )
            history_counts[table] = int(cursor.fetchone()["total"])
        if any(history_counts.values()):
            raise RuntimeError(f"AC memiliki riwayat transaksi; sinkronisasi dihentikan: {history_counts}")

        placeholders = ",".join(["%s"] * len(numbers))
        cursor.execute(
            f"SELECT nomor_inventaris,nama_aset FROM rsns_custom_logistik_non_medis_aset "
            f"WHERE nomor_inventaris IN ({placeholders}) AND TRIM(LOWER(nama_aset))<>'ac' "
            "AND NOT (kode_item='021301' AND TRIM(nama_aset) LIKE 'AC split TCL 1 PK%%' "
            "AND kode_unit IN ('61','62','63','106','107','109'))",
            numbers,
        )
        conflicts = cursor.fetchall()
        if conflicts:
            raise RuntimeError(f"Nomor AC bertabrakan dengan aset lain: {conflicts[:5]}")
    return {"history": history_counts, "numbers": len(numbers)}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--reference-workbook", type=Path, default=DEFAULT_REFERENCE_WORKBOOK)
    parser.add_argument("--host", default="localhost")
    parser.add_argument("--port", type=int, default=3306)
    parser.add_argument("--user", default="root")
    parser.add_argument("--password", default="")
    parser.add_argument("--database", default="mlite_rsns")
    parser.add_argument("--sql-output", type=Path)
    parser.add_argument("--report", type=Path)
    args = parser.parse_args()

    source_rows, excluded_rows = parse_source(args.workbook)
    price_map, year_map = load_reference_maps(args.reference_workbook)
    connection = pymysql.connect(
        host=args.host,
        port=args.port,
        user=args.user,
        password=args.password,
        database=args.database,
        charset="utf8mb4",
        autocommit=False,
        cursorclass=pymysql.cursors.DictCursor,
    )
    try:
        rows = enrich_rows(source_rows, connection, price_map, year_map)
        database_validation = validate_database(rows, connection)
    finally:
        connection.rollback()
        connection.close()

    sql = build_sql(rows)
    summary = {
        "source_groups": EXPECTED_SOURCE_GROUPS,
        "source_quantity": EXPECTED_SOURCE_QUANTITY,
        "excluded_historical_rows": [
            {"sheet": row["sheet"], "row": row["row"], "inventory": row["raw_inventory"]}
            for row in excluded_rows
        ],
        "user_addition_groups": sum(row["kind"] == "user_addition" for row in rows),
        "final_groups": len(rows),
        "final_quantity": sum(len(row["numbers"]) for row in rows),
        "unit_102_quantity": sum(len(row["numbers"]) for row in rows if row["unit"] == "102"),
        "required_room_counts": {
            code: sum(len(row["numbers"]) for row in rows if row["unit"] == code)
            for code in REQUIRED_ROOM_AC
        },
        "database_validation": database_validation,
    }

    if summary["final_groups"] != 111 or summary["final_quantity"] != 124:
        raise ValueError(f"Ringkasan akhir tidak sesuai target: {summary}")

    if args.sql_output:
        args.sql_output.parent.mkdir(parents=True, exist_ok=True)
        args.sql_output.write_text(sql, encoding="utf-8")
    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
